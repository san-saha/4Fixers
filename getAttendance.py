# -*- coding: utf-8 -*-
"""
Created on Mon Jul  5 20:37:55 2021

@author: user
"""

import pandas as pd
from datetime import datetime
import openpyxl

fmt = '%m/%d/%Y, %I:%M:%S %p'

filename=input("Enter input filename: ")#"attdn2.csv"
f = open(filename,'r',encoding='utf-16-le')
lines=[]
for i in range(5):  # skip first 11 rows
    line=f.readline()
    lines.append(line)
    
start_time_str=lines[-2].replace('"','').split("\t")[1].strip()
end_time_str=lines[-1].replace('"','').split("\t")[1].strip()

start_time_str=datetime.strptime(start_time_str,fmt)
end_time_str=datetime.strptime(end_time_str,fmt)

#print(start_time_str, type(start_time_str))
#print(end_time_str, type(end_time_str))

today=pd.read_csv(filename,sep="\t+",engine="python", header=5,encoding='utf-16-le')
today=today.sort_values(by=['Full Name'])

# for educational MS teams

fmt = '%m/%d/%Y, %I:%M:%S %p'
format_attdn={}
prv_name=''
prv_time=0
tot_time=pd.Timedelta(0)
for index,row in today.iterrows():
    #get all current values
    curr_name=row['Full Name']
    curr_duration=today.loc[index]['Duration']
    curr_duration=pd.Timedelta(curr_duration)
    
    if(curr_name==prv_name):
        tot_time=tot_time+curr_duration
    else:    
        count=0
        tot_time=curr_duration
        prv_name=curr_name
    format_attdn[curr_name]=tot_time
    
difference = end_time_str - start_time_str

xlFile=input("Enter excel filename .xlsx extension: ")#"output.xlsx"#
wb_obj = openpyxl.load_workbook(xlFile.strip())

sheet_obj = wb_obj.active

max_column=sheet_obj.max_column
max_row=sheet_obj.max_row
salary_cell=sheet_obj.cell(row=1,column=max_column+1)
date=input("Enter date: ")
salary_cell.value=date

for j in range(2, max_row):
    cell_name=sheet_obj.cell(row=j,column=1)
    salary_cell=sheet_obj.cell(row=j,column=max_column+1)
    if (format_attdn[str(cell_name.value)]):
        duration=format_attdn[str(cell_name.value)]
        percent=round(duration/difference*100)
        if(percent>100):
            percent=100
        if(percent>=60):
            salary_cell.value="P:"+str(percent)
        else:
            salary_cell.value="A:"+str(percent)
    
wb_obj.save(xlFile)
print("attendance report successfully updated excel")